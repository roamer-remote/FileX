# Copyright (c) 2026 徐泽宇
"""xlsx 业务逻辑模块。

Authors:
    徐泽宇
"""

from __future__ import annotations

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

from config import KB_EXTRACT_XLSX_MAX_COLS, KB_EXTRACT_XLSX_MAX_ROWS
from services.extract.base import ExtractResult
from services.extract.loc_markers import format_sheet_marker


def build_xlsx_marked_body(path: str) -> str:
    if load_workbook is None:
        raise ImportError("openpyxl is required to extract xlsx files")
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet_index, sheet in enumerate(wb.worksheets, start=1):
            parts.append(format_sheet_marker(sheet_index, sheet.title or ""))
            parts.append(f"## {sheet.title}")
            row_num = 0
            for row in sheet.iter_rows(max_col=KB_EXTRACT_XLSX_MAX_COLS):
                row_num += 1
                if row_num > KB_EXTRACT_XLSX_MAX_ROWS:
                    parts.append("...(truncated)")
                    break
                cells = []
                for cell in row:
                    v = cell.value
                    if v is None:
                        cells.append("")
                    else:
                        cells.append(str(v).strip())
                if any(cells):
                    parts.append(" | ".join(cells))
    finally:
        wb.close()
    return "\n".join(parts).strip()


def extract_xlsx(path: str) -> ExtractResult:
    return ExtractResult(text=build_xlsx_marked_body(path), engine="openpyxl")
